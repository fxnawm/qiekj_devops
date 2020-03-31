# coding=utf-8
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, Alignment
import pymysql
import os, sys
import oss2

class OssSdk:
    def __init__(self, path):
        self.AccessKeyId = '***'
        self.AccessKeySecret = '***'
        self.endpoint = 'http://oss-cn-shanghai.aliyuncs.com'
        self.bucketname = 'qiekj-static'
        self.path = path

    def upload(self):
        auth = oss2.Auth(self.AccessKeyId, self.AccessKeySecret)
        bucket = oss2.Bucket(auth, self.endpoint, self.bucketname)
        bucket.put_object_from_file('soft/machine_day.xlsx', self.path)


class MySQL:
    def __init__(self):
        try:
            self.con = pymysql.connect(host='qiekjodps-404f1989.cn-shanghai-1.ads.aliyuncs.com',
                                       user='***', passwd='***', db='qiekjodps',
                                       charset='utf8mb4', port=10041, )
            self.cursor = self.con.cursor()
        except Exception as e:
            msg = "\033[31;1mError:Can't connect,{}".format(e.args[1])
            print(msg)
            os._exit(0)

    def query_tuple(self, sql):
        self.cursor.execute(sql)
        return self.cursor.fetchall()

    def query_insert(self, sql):
        self.cursor.execute(sql)
        return self.cursor.fetchall()


class Excel:
    def __init__(self,select_sql):
        self.select_sql = select_sql
        self.font = Font(name='等线', size=12, bold=False, italic=False, vertAlign=None, underline='none', strike=False,
                         color='FF000000', )
        self.border = Border(left=Side(style='thin', color='FF000000'), right=Side(style='thin', color='FF000000'),
                             top=Side(style='thin', color='FF000000'), bottom=Side(style='thin', color='FF000000'))
        self.alignment = Alignment(horizontal='right', vertical='center', text_rotation=0, wrap_text=False,
                                   shrink_to_fit=False,
                                   indent=0)

    def main(self):
        client = MySQL()
        data = client.query_tuple(self.select_sql)
        for x in data:
            data = ["%.2f%%" % (i * 100) if isinstance(i, float) and i < 1 else i for i in x]
            wb = load_workbook("machine_day.xlsx")
            sheet_name = wb.sheetnames[0]
            sheet = wb[sheet_name]
            column = sheet.max_column
            row = sheet.max_row + 1
            for i in data:
                for column in range(1, len(data) + 1):
                    cell = sheet.cell(row=row, column=column)
                    cell.font = self.font
                    cell.border = self.border
                    cell.alignment = self.alignment
                    cell.value = data[column - 1]
            wb.save('machine_day.xlsx')

if __name__ == '__main__':
    # with open('merchant_yunying_data.sql',encoding='utf-8') as f:
    #     insert_sql = f.read()
    with open('machine_day.sql',encoding='utf-8') as f:
        machine_sql = f.read()
    Action = Excel(machine_sql)
    Action.main()
    ossclient = OssSdk('machine_day.xlsx')
    ossclient.upload()