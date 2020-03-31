# encoding=utf-8
import pymysql
import openpyxl
from datetime import datetime,date,timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
import smtplib

from_date = (date.today() + timedelta(days = -7)).strftime("%Y-%m-%d")
to_date = (date.today() + timedelta(days = -1)).strftime("%Y-%m-%d")
my_sender = 'service@qiekj.com'
my_password = '***'
my_receiver = ['hmq@qiekj.com','314805693@qq.com']

sql = '''SELECT
	a.sj 日期,
	a.zdds 总订单数,
	b.zfbdds 支付宝订单数,
	c.wxdds 微信订单数,
	d.yezf 余额订单数 
FROM
	(
SELECT
	DATE_FORMAT( createtime, '%Y-%m-%d' ) sj,
	count( * ) zdds 
FROM
	t_order 
WHERE
	createtime >= DATE_SUB(curdate(),interval 7 day)
	AND createtime < curdate()
	AND orderstatus = 2 
GROUP BY
	sj 
	) a
	LEFT JOIN (
SELECT
	DATE_FORMAT( createtime, '%Y-%m-%d' ) sj,
	count( * ) zfbdds 
FROM
	t_order 
WHERE
	createtime >= DATE_SUB(curdate(),interval 7 day)
	AND createtime < curdate()
	AND orderstatus = 2 
	AND paytype LIKE '1%' 
GROUP BY
	sj 
	) b ON a.sj = b.sj
	LEFT JOIN (
SELECT
	DATE_FORMAT( createtime, '%Y-%m-%d' ) sj,
	count( * ) wxdds 
FROM
	t_order 
WHERE
	createtime >= DATE_SUB(curdate(),interval 7 day)
	AND createtime < curdate()
	AND orderstatus = 2 
	AND paytype LIKE '3%' 
GROUP BY
	sj 
	) c ON a.sj = c.sj
	LEFT JOIN (
SELECT
	DATE_FORMAT( createtime, '%Y-%m-%d' ) sj,
	count( * ) yezf 
FROM
	t_order 
WHERE
	createtime >= DATE_SUB(curdate(),interval 7 day)
	AND createtime < curdate()
	AND orderstatus = 2 
	AND paytype = 2 
GROUP BY
	sj 
	) d ON a.sj = d.sj  order by a.sj'''

def get_datas(sql):
    # 一个传入SQL导出数据的函数
    # 连接数据库
    conn = pymysql.connect(host='***',
                           user='devops', passwd='apnI9KEpHf9FC68zaq4S',
                           db='***', charset='utf8mb4', port=3306)
    # 使用cursor() 方法创建一个游标对象 crusor
    cur = conn.cursor()
    # 使用execute() 方法执行SQL
    cur.execute(sql)
    # 获取所需要的数据
    datas = cur.fetchall()
    # 关闭连接
    cur.close()
    # 返回所需要的数据
    return datas

def get_fields(sql):
    # 一个传入SQL导出字段名的函数
    conn = pymysql.connect(host='***',
                           user='***', passwd='***',
                           db='***', charset='utf8mb4', port=3306)
    cur = conn.cursor()
    cur.execute(sql)
    # 获取所需要的字段名
    fields = cur.description
    cur.close()
    return fields

def get_excel(data,field,file):
    # 将数据和字段名写入Excel的函数
    # 新建一个工作簿对象
    new = openpyxl.Workbook()
    # 激活一个工作簿对象
    sheet = new.active
    # 给sheet命名
    sheet.title = '数据'
    # 将字段名循环写入excel第一行，因为字段格式列表中包含列表，每个列表的第一个元素才是字段名称
    for col in range(len(field)):
        # row代表行数，column代表列数，value代表单元格输入的值，行数和列数都是从1开始。
        _ = sheet.cell(row=1,column=col+1,value=u'%s' % field[col][0])
    # 将数据循环写入excel的每个单元格中
    for row in range(len(data)):
        for col in range(len(field)):
            # 因第一行写了字段名称，所以要从第二行开始写入
            _ = sheet.cell(row=row+2,column=col+1,value=u'%s' %data[row][col])
    # 将生成的excel保存
    newworkbook = new.save(file)
    # 返回生成的excel
    return newworkbook


def create_email(email_from,email_to,email_Subject,email_text,annex_path,annex_name):
    # 输入发件人昵称、收件人昵称、主题、正文、附件地址、附件名称生成一封邮件
    # 生成一个空的带附件的邮件实例
    message = MIMEMultipart()
    # 将正文以text的形式插入邮件中
    message.attach(MIMEText(email_text,'plain','utf-8'))
    # 生成发件人名称
    message['From'] = Header(email_from,'utf-8')
    # 生成收件人名称
    message['To'] = ','.join(email_to)
    # 生成邮件主题
    message['Subject'] = Header(email_Subject,'utf-8')
    # 读取附件的内容
    att1 = MIMEText(open(annex_path,'rb').read(),'base64','utf-8')
    att1["Content-Type"] = 'application/octet-stream'
    # 生成附件的名称
    att1["Content-Disposition"] = 'attachment; filename="qiekj_order.xlsx"'
    # 将附件内容插入邮件中
    message.attach(att1)
    # 返回邮件
    return message

def send_email(sender,password,receiver,msg):
    # 一个输入邮箱、密码、收件人、邮件内容发送邮件的函数
    try:
        # 发件人邮件中的SMTP服务器
        server = smtplib.SMTP_SSL('smtp.qiekj.com',465)
        server.ehlo()
        # 登录帐号
        server.login(sender,password)
        # 发送邮件
        server.sendmail(sender,receiver,msg.as_string())
        server.quit()
    except Exception:
        print("邮件发送失败")

if __name__ == '__main__':
    # 生成数据
    my_data = get_datas(sql)
    # 生成字段名称
    my_field = get_fields(sql)
    # 文件名称
    my_file_name = '企鹅科技'+from_date+'至'+to_date+'订单数.xlsx'
    # 文件路径
    file_path = '/data/script/sendmail_alipay/' + my_file_name
    # 生成Excel表格
    get_excel(my_data,my_field,file_path)

    my_email_from = 'service@qiekj.com'
    # 邮件标题
    my_email_Subject = '企鹅科技'+from_date+'至'+to_date+'订单数'
    # 邮件内容
    my_email_text = '企鹅科技'+from_date+'至'+to_date+'订单数'
    # 附件地址
    my_annex_path = file_path
    # 附件名称
    my_annex_name = my_file_name
    # 生成邮件
    my_msg = create_email(my_email_from,my_receiver,my_email_Subject,my_email_text,my_annex_path,my_file_name)
    # 发送邮件
    send_email(my_sender,my_password,my_receiver,my_msg)